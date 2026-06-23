import csv
from io import BytesIO, StringIO

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render

from apps.authentication.rbac import role_required
from apps.patients.models import Patient

from .services import parse_patient_filters, filtered_patients

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
except Exception:  # pragma: no cover
    reportlab_available = False
else:
    reportlab_available = True

try:
    import openpyxl  # noqa: F401
    from openpyxl.utils.dataframe import dataframe_to_rows
except Exception:  # pragma: no cover
    openpyxl_available = False
else:
    openpyxl_available = True

import datetime


@login_required
@role_required("Administrador", "Médico", "Analista")
def reports_page(request):
    """Página /reports/ — exportaciones vía endpoints existentes."""
    return render(request, "reports.html")


@login_required
@role_required('Administrador', 'Médico', 'Analista')
def export_patients_csv(request):
    """Exporta pacientes filtrados a CSV.

    Filtros soportados (opcional):
    - riesgo_enfermedad
    """
    filters = parse_patient_filters(request)
    qs = filtered_patients(filters)

    output = StringIO()
    writer = csv.writer(output)

    writer.writerow(
        [
            "id_paciente",
            "nombres",
            "apellidos",
            "edad",
            "sexo",
            "imc",
            "riesgo_enfermedad",
            "fecha_consulta",
        ]
    )

    for p in qs:
        writer.writerow(
            [
                p.id_paciente,
                p.nombres,
                p.apellidos,
                p.edad,
                p.sexo,
                p.imc,
                p.riesgo_enfermedad,
                p.fecha_consulta,
            ]
        )

    response = HttpResponse(output.getvalue(), content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="pacientes.csv"'
    return response


@login_required
@role_required('Administrador', 'Médico', 'Analista')
def export_patients_pdf(request):
    """Exporta pacientes filtrados a PDF (reportlab) con tabla formateada.

    Filtros soportados (opcional):
    - riesgo_enfermedad

    Los pacientes se ordenan por ID de menor a mayor.
    """
    if not reportlab_available:
        return HttpResponse("reportlab no está instalado en el entorno.", status=500)

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    )

    filters = parse_patient_filters(request)
    qs = filtered_patients(filters).order_by("id_paciente")  # Ordenado por ID ascendente

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleStyle",
        parent=styles["Heading1"],
        fontSize=16,
        spaceAfter=12,
        textColor=colors.HexColor("#0d6efd"),
    )
    subtitle_style = ParagraphStyle(
        "SubtitleStyle",
        parent=styles["Normal"],
        fontSize=10,
        spaceAfter=20,
        textColor=colors.gray,
    )

    elements = []

    # Título
    elements.append(Paragraph("HealthAnalytics IPS", title_style))
    filtro_texto = f"Filtro: {filters.riesgo_enfermedad}" if filters.riesgo_enfermedad else "Filtro: Todos los pacientes"
    elements.append(Paragraph(
        f"Reporte de Pacientes — {filtro_texto} — {datetime.date.today().isoformat()}",
        subtitle_style,
    ))

    total = qs.count()
    elements.append(Paragraph(f"Total de pacientes: <b>{total}</b>", styles["Normal"]))
    elements.append(Spacer(1, 12))

    # Datos de la tabla
    header = [
        "ID", "Nombres", "Apellidos", "Edad", "Sexo",
        "IMC", "Pres. Sist.", "Pres. Diast.", "Glucosa",
        "Colesterol", "Riesgo", "Fecha",
    ]

    data = [header]
    for p in qs[:200]:
        data.append([
            str(p.id_paciente),
            p.nombres or "",
            p.apellidos or "",
            str(p.edad),
            p.sexo or "",
            f"{p.imc:.1f}" if p.imc else "",
            str(p.presion_sistolica),
            str(p.presion_diastolica),
            f"{p.glucosa:.0f}" if p.glucosa else "",
            f"{p.colesterol:.0f}" if p.colesterol else "",
            p.riesgo_enfermedad or "",
            str(p.fecha_consulta) if p.fecha_consulta else "",
        ])

    # Calcular anchos de columna proporcionales
    col_widths = [30, 60, 60, 30, 40, 35, 40, 40, 40, 45, 50, 60]
    available_width = letter[0] - 60  # márgenes
    scale = available_width / sum(col_widths)
    col_widths = [w * scale for w in col_widths]

    table = Table(data, colWidths=col_widths, repeatRows=1)

    # Estilo de tabla profesional
    table_style = TableStyle([
        # Encabezado
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0d6efd")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("TOPPADDING", (0, 0), (-1, 0), 8),
        # Filas de datos
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
        ("TOPPADDING", (0, 1), (-1, -1), 5),
        # Alineación
        ("ALIGN", (0, 0), (0, -1), "CENTER"),   # ID centrado
        ("ALIGN", (3, 0), (3, -1), "CENTER"),   # Edad centrado
        ("ALIGN", (5, 0), (10, -1), "CENTER"),  # IMC, presiones, glucosa, colesterol, riesgo centrados
        # Líneas de cuadrícula
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dee2e6")),
        ("LINEBELOW", (0, 0), (-1, 0), 1.5, colors.HexColor("#0a58ca")),
        # Colores alternados
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
    ])

    # Colorear la columna de riesgo
    riesgo_colors = {
        "Crítico": colors.HexColor("#dc3545"),
        "Alto": colors.HexColor("#fd7e14"),
        "Medio": colors.HexColor("#0dcaf0"),
        "Bajo": colors.HexColor("#198754"),
    }
    for row_idx, row_data in enumerate(data[1:], start=1):
        riesgo = row_data[10]
        if riesgo in riesgo_colors:
            table_style.add("TEXTCOLOR", (10, row_idx), (10, row_idx), riesgo_colors[riesgo])
            table_style.add("FONTNAME", (10, row_idx), (10, row_idx), "Helvetica-Bold")

    table.setStyle(table_style)
    elements.append(table)

    # Pie de página
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(
        f"Documento generado el {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')} — HealthAnalytics IPS",
        styles["Normal"],
    ))

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="pacientes.pdf"'
    return response


@login_required
@role_required('Administrador', 'Médico', 'Analista')
def export_patients_excel(request):
    """Exporta Excel (.xlsx) con 5 hojas usando la MISMA fuente base que CSV/PDF.

    Requiere openpyxl.

    Filtros soportados (opcional):
    - riesgo_enfermedad
    """
    if not openpyxl_available:
        return HttpResponse("openpyxl no está instalado en el entorno.", status=500)

    filters = parse_patient_filters(request)
    qs = filtered_patients(filters)

    # Import local para no romper import si no se usa.
    from openpyxl import Workbook

    wb = Workbook()

    # Hoja 1: Pacientes
    ws_pac = wb.active
    ws_pac.title = "Pacientes"
    headers = [
        "id_paciente",
        "nombres",
        "apellidos",
        "edad",
        "sexo",
        "imc",
        "riesgo_enfermedad",
        "fecha_consulta",
    ]
    ws_pac.append(headers)
    for p in qs:
        ws_pac.append(
            [
                p.id_paciente,
                p.nombres,
                p.apellidos,
                p.edad,
                p.sexo,
                p.imc,
                p.riesgo_enfermedad,
                p.fecha_consulta,
            ]
        )

    # Hoja 2: KPIs (recalculados desde la misma query base)
    ws_kpi = wb.create_sheet("KPIs")
    ws_kpi.append(["kpi", "valor"])
    # KPI simples para evitar dependencia en analytics.services
    risco_count = {
        "Bajo": qs.filter(riesgo_enfermedad="Bajo").count(),
        "Medio": qs.filter(riesgo_enfermedad="Medio").count(),
        "Alto": qs.filter(riesgo_enfermedad="Alto").count(),
        "Crítico": qs.filter(riesgo_enfermedad="Crítico").count(),
    }
    ws_kpi.append(["total_pacientes", qs.count()])
    for k, v in risco_count.items():
        ws_kpi.append([f"pacientes_{k.lower()}", v])

    # Hoja 3: Segmentaciones (básico por edad bucket e IMC bucket)
    ws_seg = wb.create_sheet("Segmentaciones")
    ws_seg.append(["segmento", "valor"])
    # edad buckets
    def age_bucket(age: int) -> str:
        if age < 18:
            return "<18"
        if age < 35:
            return "18-34"
        if age < 50:
            return "35-49"
        if age < 65:
            return "50-64"
        return "65+"

    age_counts = {}
    for age in qs.values_list("edad", flat=True):
        key = age_bucket(int(age))
        age_counts[key] = age_counts.get(key, 0) + 1

    for k, v in sorted(age_counts.items()):
        ws_seg.append([f"edad:{k}", v])

    # Hoja 4: Predicciones (no persistidas)
    ws_pred = wb.create_sheet("Predicciones")
    ws_pred.append(["id_paciente", "riesgo_enfermedad", "confidence"])
    # Para no romper si el modelo no existe, se llenan solo si hay.
    from apps.ml.predict import predict_risk

    for p in qs.order_by("-fecha_consulta")[:50]:
        try:
            risk_label, confidence, _ = predict_risk(
                {
                    "edad": p.edad,
                    "imc": p.imc,
                    "presion_sistolica": p.presion_sistolica,
                    "presion_diastolica": p.presion_diastolica,
                    "glucosa": p.glucosa,
                    "colesterol": p.colesterol,
                }
            )
            ws_pred.append([p.id_paciente, str(risk_label), float(confidence)])
        except Exception:
            continue

    # Hoja 5: Historial ETL
    from apps.etl.models import ETLRun

    ws_etl = wb.create_sheet("Historial ETL")
    ws_etl.append(["started_at", "records_processed", "status"])
    for r in ETLRun.objects.all().order_by("-started_at")[:50]:
        ws_etl.append([r.started_at.isoformat(), int(r.records_processed), r.status])

    # Metadata
    ws_meta = wb.create_sheet("Meta")
    ws_meta.append(["generated_at", datetime.datetime.now().isoformat()])
    ws_meta.append(["filter_riesgo_enfermedad", filters.riesgo_enfermedad or "(none)"])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="healthanalytics_report.xlsx"'
    return response

